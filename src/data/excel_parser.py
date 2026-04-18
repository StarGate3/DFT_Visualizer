"""Excel (.xlsx) parser for DFT calculation results."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from src.data.models import (
    CompoundFranckCondon,
    CompoundHomoLumo,
    CompoundStates,
    DFTDataset,
)


class ExcelParsingError(Exception):
    """Raised when the Excel file cannot be parsed correctly."""


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _get_header_map(ws: Worksheet, expected_cols: list[str]) -> dict[str, int]:
    """Locate the header row and return a column-name → 0-based-index mapping.

    Searches rows from the top until it finds one whose first non-None cell
    matches the first expected column name (case-insensitive).

    Args:
        ws: Worksheet to search.
        expected_cols: Lower-cased column names that must be present.

    Returns:
        Dict mapping lower-cased column name to 0-based index.

    Raises:
        ExcelParsingError: If no header row is found or required columns are absent.
    """
    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if first is None:
            continue
        if str(first).strip().lower() != expected_cols[0].lower():
            continue
        header_map: dict[str, int] = {}
        for idx, cell in enumerate(row):
            if cell is not None:
                header_map[str(cell).strip().lower()] = idx
        missing = [c for c in expected_cols if c.lower() not in header_map]
        if missing:
            raise ExcelParsingError(
                f"Missing required columns {missing} in sheet '{ws.title}'"
            )
        return header_map
    raise ExcelParsingError(
        f"Could not find header row with column '{expected_cols[0]}' "
        f"in sheet '{ws.title}'"
    )


def _require_float(value: Any, col_name: str, row_num: int) -> float:
    """Convert *value* to float or raise ExcelParsingError with a row reference.

    Args:
        value: Raw cell value from openpyxl.
        col_name: Human-readable column name used in error messages.
        row_num: 1-based row number used in error messages.

    Returns:
        The numeric value as float.

    Raises:
        ExcelParsingError: If the value cannot be converted to float.
    """
    try:
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        raise ExcelParsingError(
            f"Invalid {col_name} value in row {row_num}: {value!r}"
        )


def _is_header_or_empty(name_cell: Any, expected_name_col: str) -> bool:
    """Return True if the row should be skipped (empty or is the header row itself)."""
    if name_cell is None:
        return True
    text = str(name_cell).strip()
    return text == "" or text.lower() == expected_name_col.lower()


# ---------------------------------------------------------------------------
# Sheet-level parsers
# ---------------------------------------------------------------------------


def _parse_homo_lumo(ws: Worksheet) -> list[CompoundHomoLumo]:
    """Parse the HOMO_LUMO worksheet into a list of CompoundHomoLumo objects."""
    expected = ["compound_name", "homo_ev", "lumo_ev"]
    header = _get_header_map(ws, expected)
    results: list[CompoundHomoLumo] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        name_cell = row[header["compound_name"]]
        if _is_header_or_empty(name_cell, "compound_name"):
            continue
        name = str(name_cell).strip()
        homo = _require_float(row[header["homo_ev"]], "HOMO", row_idx)
        lumo = _require_float(row[header["lumo_ev"]], "LUMO", row_idx)
        results.append(CompoundHomoLumo(name=name, homo=homo, lumo=lumo))
    return results


def _parse_states(ws: Worksheet) -> list[CompoundStates]:
    """Parse the States worksheet into a list of CompoundStates objects."""
    expected = ["compound_name", "s0_ev", "s1_ev", "t1_ev"]
    header = _get_header_map(ws, expected)
    results: list[CompoundStates] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        name_cell = row[header["compound_name"]]
        if _is_header_or_empty(name_cell, "compound_name"):
            continue
        name = str(name_cell).strip()
        s0 = _require_float(row[header["s0_ev"]], "S0", row_idx)
        s1 = _require_float(row[header["s1_ev"]], "S1", row_idx)
        t1 = _require_float(row[header["t1_ev"]], "T1", row_idx)
        results.append(CompoundStates(name=name, s0=s0, s1=s1, t1=t1))
    return results


def _parse_franck_condon(ws: Worksheet) -> list[CompoundFranckCondon]:
    """Parse the FranckCondon worksheet into a list of CompoundFranckCondon objects."""
    expected = ["compound_name", "state", "e_vertical", "e_adiabatic"]
    header = _get_header_map(ws, expected)
    results: list[CompoundFranckCondon] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        name_cell = row[header["compound_name"]]
        if _is_header_or_empty(name_cell, "compound_name"):
            continue
        name = str(name_cell).strip()
        state_cell = row[header["state"]]
        state = str(state_cell).strip() if state_cell is not None else ""
        e_vert = _require_float(row[header["e_vertical"]], "E_vertical", row_idx)
        e_adiab = _require_float(row[header["e_adiabatic"]], "E_adiabatic", row_idx)

        bde_value: float | None = None
        bde_label: str | None = None

        if "bde_value" in header:
            raw = row[header["bde_value"]]
            if raw is not None and str(raw).strip() != "":
                bde_value = _require_float(raw, "BDE_value", row_idx)

        if "bde_label" in header:
            raw_label = row[header["bde_label"]]
            if raw_label is not None and str(raw_label).strip() != "":
                bde_label = str(raw_label).strip()

        results.append(CompoundFranckCondon(
            name=name,
            state=state,
            e_vertical=e_vert,
            e_adiabatic=e_adiab,
            bde_value=bde_value,
            bde_label=bde_label,
        ))
    return results


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def parse_excel(filepath: Path) -> DFTDataset:
    """Parse a DFT results Excel file into a DFTDataset.

    Expected sheets: ``HOMO_LUMO``, ``States`` (required), ``FranckCondon``
    (optional).  Any other sheets (e.g. ``README``) are ignored.

    Args:
        filepath: Path to the ``.xlsx`` file.

    Returns:
        A fully populated :class:`DFTDataset`.

    Raises:
        ExcelParsingError: If the file cannot be opened, required sheets or
            columns are missing, or any numeric cell contains a non-numeric
            value.
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as exc:
        raise ExcelParsingError(
            f"Cannot open file '{filepath.name}': {exc}"
        ) from exc

    for required in ("HOMO_LUMO", "States"):
        if required not in wb.sheetnames:
            raise ExcelParsingError(
                f"Missing required sheet '{required}'"
            )

    homo_lumo = _parse_homo_lumo(wb["HOMO_LUMO"])
    states = _parse_states(wb["States"])
    franck_condon: list[CompoundFranckCondon] = (
        _parse_franck_condon(wb["FranckCondon"])
        if "FranckCondon" in wb.sheetnames
        else []
    )
    return DFTDataset(homo_lumo=homo_lumo, states=states, franck_condon=franck_condon)
