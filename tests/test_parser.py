"""Tests for src/data/excel_parser and src/data/models."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.data.excel_parser import ExcelParsingError, parse_excel
from src.data.models import DFTDataset

SAMPLE_XLSX = Path(__file__).parent.parent / "sample_compounds.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_minimal_xlsx(
    tmp_path: Path,
    include_franck_condon: bool = True,
    bad_homo: bool = False,
) -> Path:
    """Create a minimal valid (or intentionally broken) .xlsx for testing."""
    wb = openpyxl.Workbook()
    ws_hl = wb.active
    ws_hl.title = "HOMO_LUMO"
    ws_hl.append(["compound_name", "HOMO_eV", "LUMO_eV"])
    ws_hl.append(["TestComp", "abc" if bad_homo else -7.0, -2.0])

    ws_st = wb.create_sheet("States")
    ws_st.append(["compound_name", "S0_eV", "S1_eV", "T1_eV"])
    ws_st.append(["TestComp", 0.0, 3.0, 2.5])

    if include_franck_condon:
        ws_fc = wb.create_sheet("FranckCondon")
        ws_fc.append(["compound_name", "state", "E_vertical", "E_adiabatic",
                       "BDE_value", "BDE_label"])
        ws_fc.append(["TestComp", "S0", 0.0, 0.0, None, None])

    path = tmp_path / "test.xlsx"
    wb.save(str(path))
    return path


# ---------------------------------------------------------------------------
# Sample file tests
# ---------------------------------------------------------------------------


def test_sample_loads_successfully() -> None:
    """parse_excel must not raise on the bundled sample file."""
    dataset = parse_excel(SAMPLE_XLSX)
    assert isinstance(dataset, DFTDataset)


def test_sample_homo_lumo_count() -> None:
    """Sample file contains exactly 8 HOMO/LUMO entries."""
    dataset = parse_excel(SAMPLE_XLSX)
    assert len(dataset.homo_lumo) == 8


def test_sample_states_count() -> None:
    """Sample file contains exactly 5 states entries."""
    dataset = parse_excel(SAMPLE_XLSX)
    assert len(dataset.states) == 5


def test_benzofenon_homo_value() -> None:
    """Benzofenon HOMO must parse to -7.16 eV."""
    dataset = parse_excel(SAMPLE_XLSX)
    match = next(c for c in dataset.homo_lumo if c.name == "Benzofenon")
    assert match.homo == pytest.approx(-7.16)


def test_benzofenon_gap() -> None:
    """Benzofenon gap must equal LUMO - HOMO."""
    dataset = parse_excel(SAMPLE_XLSX)
    match = next(c for c in dataset.homo_lumo if c.name == "Benzofenon")
    assert match.gap == pytest.approx(match.lumo - match.homo)


def test_franck_condon_entries_loaded() -> None:
    """Sample FranckCondon sheet should produce at least one entry."""
    dataset = parse_excel(SAMPLE_XLSX)
    assert len(dataset.franck_condon) > 0


def test_franck_condon_bde_none_preserved() -> None:
    """Rows with empty BDE cells must store None, not a string."""
    dataset = parse_excel(SAMPLE_XLSX)
    none_entries = [e for e in dataset.franck_condon if e.bde_value is None]
    assert none_entries, "Expected at least one FC entry with no BDE value"


# ---------------------------------------------------------------------------
# Missing FranckCondon sheet
# ---------------------------------------------------------------------------


def test_missing_franck_condon_sheet_ok(tmp_path: Path) -> None:
    """A file without FranckCondon sheet must succeed with empty fc list."""
    path = _make_minimal_xlsx(tmp_path, include_franck_condon=False)
    dataset = parse_excel(path)
    assert dataset.franck_condon == []


# ---------------------------------------------------------------------------
# Error handling
# ---------------------------------------------------------------------------


def test_nonexistent_file_raises_error(tmp_path: Path) -> None:
    """Passing a non-existent path must raise ExcelParsingError."""
    with pytest.raises(ExcelParsingError):
        parse_excel(tmp_path / "does_not_exist.xlsx")


def test_corrupted_file_raises_error(tmp_path: Path) -> None:
    """A file that is not a valid xlsx must raise ExcelParsingError."""
    bad = tmp_path / "bad.xlsx"
    bad.write_bytes(b"this is not a zip archive")
    with pytest.raises(ExcelParsingError):
        parse_excel(bad)


def test_missing_required_sheet_raises_error(tmp_path: Path) -> None:
    """A file missing the HOMO_LUMO sheet must raise ExcelParsingError."""
    wb = openpyxl.Workbook()
    wb.active.title = "States"  # type: ignore[union-attr]
    wb["States"].append(["compound_name", "S0_eV", "S1_eV", "T1_eV"])
    path = tmp_path / "no_hl.xlsx"
    wb.save(str(path))
    with pytest.raises(ExcelParsingError, match="HOMO_LUMO"):
        parse_excel(path)


def test_invalid_numeric_value_raises_error(tmp_path: Path) -> None:
    """Non-numeric HOMO cell must raise ExcelParsingError with row info."""
    path = _make_minimal_xlsx(tmp_path, bad_homo=True)
    with pytest.raises(ExcelParsingError, match="HOMO"):
        parse_excel(path)


def test_readme_sheet_ignored(tmp_path: Path) -> None:
    """Extra sheets like README must be silently ignored."""
    path = _make_minimal_xlsx(tmp_path)
    wb = openpyxl.load_workbook(str(path))
    readme = wb.create_sheet("README")
    readme.append(["Some documentation text"])
    wb.save(str(path))
    dataset = parse_excel(path)
    assert len(dataset.homo_lumo) == 1
